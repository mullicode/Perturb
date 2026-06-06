import argparse
import datetime
import logging as pylogging
import os
import threading
import time
import typing
import math
import asyncio

import bittensor as bt
import torch
import torch.nn.functional as F

from perturbnet.image_io import decode_image_b64, encode_image_b64
from openpyxl import Workbook, load_workbook

from perturbnet.model import (
    _preprocess_for_efficientnet_v2_l,
    load_efficientnet_v2_l,
    logits_for_images,
    resolve_target_index,
)
from perturbnet.protocol import AttackChallenge
from perturbnet import constants as C

logger = pylogging.getLogger(__name__)

_ATTACK_EXCEL_LOCK = threading.Lock()
_ATTACK_EXCEL_HEADERS = (
    "timestamp",
    "task_id",
    "model_name",
    "prompt",
    "true_label",
    "epsilon",
    "norm_type",
    "min_delta",
    "resolution",
    "prediction",
    "progress",
    "rmse",
    "norm",
    "estimated_score",
)


def _append_attack_excel_row(
    excel_path: str,
    *,
    task_id: str,
    model_name: str,
    prompt: str,
    true_label: str,
    epsilon: float,
    norm_type: str,
    min_delta: float,
    resolution: str,
    progress: str = "",
    prediction: typing.Optional[int] = None,
    rmse: typing.Optional[float] = None,
    norm: typing.Optional[float] = None,
    estimated_score: typing.Optional[float] = None,
) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(excel_path)) or ".", exist_ok=True)
    row = [
        datetime.datetime.now().isoformat(timespec="seconds"),
        task_id,
        model_name,
        prompt,
        true_label,
        epsilon,
        norm_type,
        min_delta,
        resolution,
        "" if prediction is None else prediction,
        progress,
        "" if rmse is None else rmse,
        "" if norm is None else norm,
        "" if estimated_score is None else estimated_score,
    ]
    with _ATTACK_EXCEL_LOCK:
        if os.path.isfile(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "attack_log"
            ws.append(list(_ATTACK_EXCEL_HEADERS))
        ws.append(row)
        wb.save(excel_path)


class _AttackExcelRecorder:
    """Accumulates attack progress and writes one Excel row per task."""

    def __init__(
        self,
        excel_path: str,
        *,
        task_id: str,
        model_name: str,
        prompt: str,
        true_label: str,
        epsilon: float,
        norm_type: str,
        min_delta: float,
        resolution: str,
    ) -> None:
        self._excel_path = excel_path
        self._task_id = task_id
        self._model_name = model_name
        self._prompt = prompt
        self._true_label = true_label
        self._epsilon = epsilon
        self._norm_type = norm_type
        self._min_delta = min_delta
        self._resolution = resolution
        self._progress: list[str] = []
        self._prediction: typing.Optional[int] = None
        self._rmse: typing.Optional[float] = None
        self._norm: typing.Optional[float] = None
        self._estimated_score: typing.Optional[float] = None

    def log(
        self,
        progress: str,
        prediction: typing.Optional[int] = None,
        rmse: typing.Optional[float] = None,
        norm: typing.Optional[float] = None,
        estimated_score: typing.Optional[float] = None,
    ) -> None:
        self._progress.append(progress)
        if prediction is not None:
            self._prediction = prediction
        if rmse is not None:
            self._rmse = rmse
        if norm is not None:
            self._norm = norm
        if estimated_score is not None:
            self._estimated_score = estimated_score

    def set_resolution(self, resolution: str) -> None:
        self._resolution = resolution

    def flush(self) -> None:
        if not self._progress:
            return
        _append_attack_excel_row(
            self._excel_path,
            task_id=self._task_id,
            model_name=self._model_name,
            prompt=self._prompt,
            true_label=self._true_label,
            epsilon=self._epsilon,
            norm_type=self._norm_type,
            min_delta=self._min_delta,
            resolution=self._resolution,
            progress=" | ".join(self._progress),
            prediction=self._prediction,
            rmse=self._rmse,
            norm=self._norm,
            estimated_score=self._estimated_score,
        )


# ── Validator acceptance window ─────────────────────────────────────────────
# These MUST mirror neurons/validator.py::verify_and_score + the canonical
# perturbnet/constants.py defaults. They are VALIDITY gates (what the validator
# will actually accept). The validator's L-inf window is [min_linf, min(epsilon,
# max_linf)] = [0.003, 0.03] for the sampled epsilon range [0.06, 0.20], and it
# never rejects on RMSE (RMSE only affects the score).
_VAL_MIN_LINF      = float(os.getenv("MINER_VAL_MIN_LINF",    "0.003"))
_VAL_MAX_LINF      = float(os.getenv("MINER_VAL_MAX_LINF",    "0.00393"))
_VAL_MIN_SSIM      = float(os.getenv("MINER_VAL_MIN_SSIM",    "0.98"))
_VAL_MIN_PSNR_DB   = float(os.getenv("MINER_VAL_MIN_PSNR_DB", "38.0"))

# ── Trust-region / marginal-gain batch controller ─────────────────────────
# Batch size is selected by real measured gap reduction, not only gap / gap_rate.
_MARGINAL_PROBE_ENABLED = os.getenv("MINER_MARGINAL_PROBE_ENABLED", "1").strip() not in ("0", "false", "no")

# Prefer smaller batch if a larger batch is not meaningfully better.
_BATCH_EFFICIENCY_TOL = float(os.getenv("MINER_BATCH_EFFICIENCY_TOL", "0.88"))

# Safety-grow should not use very large batches because it only needs margin buffer.
_MAX_SAFETY_GROW_BATCH = int(os.getenv("MINER_MAX_SAFETY_GROW_BATCH", "4"))
# ── Adaptive doubling + trust-region refinement ───────────────────────────
# Coarse doubling finds the useful batch region.
# Refinement checks inside the good/bad interval, e.g. 512 good, 1024 weak
# then test 640/768/896.
_MAX_COARSE_PROBES = int(os.getenv("MINER_MAX_COARSE_PROBES", "5"))
_MAX_REFINE_PROBES = int(os.getenv("MINER_MAX_REFINE_PROBES", "9"))
_REFINE_TIME_MIN = float(os.getenv("MINER_REFINE_TIME_MIN", "0.75"))

# Hard expansion cap. Stage limits are only the first trust-region range.
# If the largest candidate is still efficient, the controller may expand upward.
_MAX_EXPANDED_BATCH = int(os.getenv("MINER_MAX_EXPANDED_BATCH", "4096"))

# If added pixels after a larger candidate give less than this ratio of the
# previous candidate efficiency, treat the larger candidate as weak.
_TRUST_REGION_COLLAPSE_RATIO = float(os.getenv("MINER_TRUST_REGION_COLLAPSE_RATIO", "0.45"))

# Require meaningful absolute progress so the controller does not greedily pick
# very tiny efficient batches forever.
_MIN_PROGRESS_FAR = float(os.getenv("MINER_MIN_PROGRESS_FAR", "0.15"))
_MIN_PROGRESS_MID = float(os.getenv("MINER_MIN_PROGRESS_MID", "0.20"))
_MIN_PROGRESS_NEAR = float(os.getenv("MINER_MIN_PROGRESS_NEAR", "0.30"))
_FLIP_MARGIN_EPS      = float(os.getenv("MINER_FLIP_MARGIN_EPS", "-0.0005"))
_STRONG_SAFE_MARGIN   = float(os.getenv("MINER_STRONG_SAFE_MARGIN", "-0.0015"))
_SAFETY_GROW_TARGET   = float(os.getenv("MINER_SAFETY_GROW_TARGET", "-0.0010"))
_NEAR_BOUNDARY_GAP    = float(os.getenv("MINER_NEAR_BOUNDARY_GAP", "0.03"))
_TARGET_SWITCH_STEPS  = int(os.getenv("MINER_TARGET_SWITCH_STEPS", "8"))
# Match the validator's numerics by disabling TF32 reductions on the miner so
# a locally-confirmed flip reflects what the validator will actually compute.
_DISABLE_TF32      = os.getenv("MINER_DISABLE_TF32", "1").strip() not in ("0", "false", "no")
if _DISABLE_TF32:
    try:
        torch.backends.cuda.matmul.allow_tf32 = False
        torch.backends.cudnn.allow_tf32 = False
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════
#  Model forward helpers
# ═══════════════════════════════════════════════════════════════════════════

def _model_logits_batch(model: torch.nn.Module, x_in: torch.Tensor) -> torch.Tensor:
    x_prep = _preprocess_for_efficientnet_v2_l(x_in)
    return model(x_prep)


def _pred_index(model: torch.nn.Module, image_chw: torch.Tensor) -> int:
    with torch.inference_mode():
        logits = _model_logits_batch(model, image_chw.unsqueeze(0).float())
        return int(logits.argmax(dim=1)[0].item())


# ═══════════════════════════════════════════════════════════════════════════
#  Bittensor plumbing
# ═══════════════════════════════════════════════════════════════════════════

def _make_wallet(config):
    wallet_name = getattr(config.wallet, "name", getattr(config, "wallet_name", "default"))
    wallet_hotkey = getattr(config.wallet, "hotkey", getattr(config, "wallet_hotkey", "default"))
    if hasattr(bt, "wallet"):
        try:
            return bt.wallet(name=wallet_name, hotkey=wallet_hotkey)
        except Exception:
            return bt.wallet(config=config)
    wallet_cls = getattr(bt, "Wallet", None)
    if wallet_cls is None:
        raise RuntimeError("No wallet constructor found in bittensor.")
    try:
        return wallet_cls(name=wallet_name, hotkey=wallet_hotkey)
    except TypeError:
        return wallet_cls(config=config)


def _make_subtensor(config):
    network = getattr(config.subtensor, "network", getattr(config, "network", "finney"))
    chain_endpoint = getattr(config.subtensor, "chain_endpoint", None) or getattr(config, "chain_endpoint", None)
    if hasattr(bt, "subtensor"):
        if chain_endpoint:
            try:
                return bt.subtensor(chain_endpoint=chain_endpoint)
            except Exception:
                pass
        try:
            return bt.subtensor(network=network)
        except Exception:
            return bt.subtensor(config=config)
    subtensor_cls = getattr(bt, "Subtensor", None)
    if subtensor_cls is None:
        raise RuntimeError("No subtensor constructor found in bittensor.")
    if chain_endpoint:
        try:
            return subtensor_cls(chain_endpoint=chain_endpoint)
        except Exception:
            pass
    try:
        return subtensor_cls(network=network)
    except Exception:
        return subtensor_cls(config=config)


def _make_axon(wallet, config) -> typing.Any:
    axon_cfg = getattr(config, "axon", None)
    port     = int(
        getattr(axon_cfg, "port", None)
        or getattr(config, "axon_port", None)
        or os.getenv("MINER_PORT", os.getenv("AXON_PORT", "9000"))
    )
    ip = str(
        getattr(axon_cfg, "ip", None) or os.getenv("MINER_IP", os.getenv("AXON_IP", "0.0.0.0"))
    ).strip() or "0.0.0.0"
    external_ip = str(
        getattr(axon_cfg, "external_ip", None) or os.getenv("MINER_EXTERNAL_IP", "")
    ).strip()
    external_port_raw = (
        getattr(axon_cfg, "external_port", None) or os.getenv("MINER_EXTERNAL_PORT", "")
    )
    external_port = int(str(external_port_raw).strip()) if str(external_port_raw).strip() else port
    if not external_ip:
        raise RuntimeError(
            "MINER_EXTERNAL_IP is not set. "
            "Set it to the public IP address that validators can reach this miner on."
        )
    max_workers = int(getattr(axon_cfg, "max_workers", None) or os.getenv("AXON_MAX_WORKERS", "10"))
    axon_cls    = getattr(bt, "Axon", None)
    if axon_cls is None:
        raise RuntimeError("bittensor.Axon class not found.")
    logger.info(
        f"[MINER] Creating axon ip={ip} port={port} "
        f"external_ip={external_ip} external_port={external_port} max_workers={max_workers}"
    )
    return axon_cls(
        wallet=wallet, ip=ip, port=port,
        external_ip=external_ip, external_port=external_port,
        max_workers=max_workers,
    )


def _configure_log_level(level_raw: str) -> None:
    level_name = (level_raw or "DEBUG").upper()
    requested_level = getattr(pylogging, level_name, pylogging.INFO)
    level = max(int(pylogging.INFO), int(requested_level))
    pylogging.basicConfig(
        level=level,
        format="%(asctime)s | %(name)s | %(levelname)s | %(message)s",
    )
    pylogging.getLogger().setLevel(level)

# ═══════════════════════════════════════════════════════════════════════════
#  Quality / preflight / PNG submit
# ═══════════════════════════════════════════════════════════════════════════

class _AdvQuality(typing.NamedTuple):
    norm:       float
    rmse:       float
    ssim:       float
    psnr_db:    float
    pred:       int
    target_hit: bool
    flipped:    bool
    # logits[true] - max_{j != true} logits[j].  < 0 ⇒ argmax flipped;
    # <= -_FLIP_MARGIN_EPS ⇒ robust (transferable) flip.
    margin:     float = 0.0


class _PreflightResult(typing.NamedTuple):
    ok:      bool
    reason:  str
    quality: _AdvQuality


def _compute_ssim(x_clean: torch.Tensor, x_adv: torch.Tensor, kernel_size: int = 11) -> float:
    if x_clean.ndim != 3 or x_adv.ndim != 3 or x_clean.shape != x_adv.shape:
        return 0.0
    padding = kernel_size // 2
    x, y    = x_clean.unsqueeze(0), x_adv.unsqueeze(0)
    c1, c2  = 0.01 ** 2, 0.03 ** 2
    mu_x    = F.avg_pool2d(x, kernel_size=kernel_size, stride=1, padding=padding)
    mu_y    = F.avg_pool2d(y, kernel_size=kernel_size, stride=1, padding=padding)
    sigma_x  = F.avg_pool2d(x * x, kernel_size=kernel_size, stride=1, padding=padding) - mu_x * mu_x
    sigma_y  = F.avg_pool2d(y * y, kernel_size=kernel_size, stride=1, padding=padding) - mu_y * mu_y
    sigma_xy = F.avg_pool2d(x * y, kernel_size=kernel_size, stride=1, padding=padding) - mu_x * mu_y
    numerator   = (2.0 * mu_x * mu_y + c1) * (2.0 * sigma_xy + c2)
    denominator = (mu_x * mu_x + mu_y * mu_y + c1) * (sigma_x + sigma_y + c2)
    return float((numerator / (denominator + 1e-12)).mean().item())


def _compute_psnr_db(x_clean: torch.Tensor, x_adv: torch.Tensor) -> float:
    mse = float(torch.mean((x_adv - x_clean) ** 2).item())
    if mse <= 1e-12:
        return 99.0
    return 10.0 * math.log10(1.0 / mse)


def _measure_adv_quality(
    model: torch.nn.Module,
    clean: torch.Tensor,
    adv: torch.Tensor,
    true_label: int,
    target_label: int,
) -> _AdvQuality:
    diff = adv - clean
    norm = float(diff.abs().max().item())
    rmse = float(torch.sqrt(torch.mean(diff ** 2)).item())
    tl = int(true_label)
    with torch.inference_mode():
        logits = _model_logits_batch(model, adv.unsqueeze(0).float()).squeeze(0).float()
    pred = int(logits.argmax().item())
    # margin = logits[true] - best competing logit.  A robust flip needs the
    # true label to sit at least _FLIP_MARGIN_EPS below the winner, otherwise
    # the boundary flip will not transfer to the validator's inference path.
    competitor = logits.clone()
    competitor[tl] = float("-inf")
    margin = float(logits[tl].item() - competitor.max().item())
    robust_flip = (pred != tl) and (margin <= _FLIP_MARGIN_EPS)
    return _AdvQuality(
        norm=norm, rmse=rmse,
        ssim=_compute_ssim(clean, adv), psnr_db=_compute_psnr_db(clean, adv),
        pred=pred, target_hit=pred == int(target_label), flipped=robust_flip,
        margin=margin,
    )


def _encode_decode_roundtrip(adv: torch.Tensor) -> torch.Tensor:
    """Single PNG encode then decode — matches exactly what validator does."""
    return decode_image_b64(encode_image_b64(adv)).to(adv.device)


def _quality_on_png(
    model: torch.nn.Module,
    clean: torch.Tensor,
    adv: torch.Tensor,
    true_label: int,
    target_label: int,
) -> typing.Tuple[torch.Tensor, _AdvQuality]:
    """Measure quality against the original clean image via one PNG roundtrip."""
    decoded = _encode_decode_roundtrip(adv)
    q = _measure_adv_quality(model, clean, decoded, true_label, target_label)
    return decoded, q


def _preflight_flip_only(
    quality: _AdvQuality,
    true_label: int,
    epsilon: float,
) -> _PreflightResult:
    tl = int(true_label)
    # Mirror neurons/validator.py::verify_and_score acceptance gates exactly.
    # norm/rmse/ssim/psnr are already measured the same way the validator
    # measures them (both clean and adv decoded from PNG), so these checks
    # predict the validator outcome faithfully. The validator's L-inf ceiling
    # is min(epsilon, max_linf_delta) — NOT the miner's 1/255 optimization
    # target — and it has no RMSE rejection.
    eff_max = min(float(epsilon), _VAL_MAX_LINF)
    if quality.pred == tl:
        return _PreflightResult(False, "label_match_with_original", quality)
    # Hard flip gate: margin must be safely below zero to survive validator numerics.
    if quality.margin > _FLIP_MARGIN_EPS:
        return _PreflightResult(False, "weak_flip_margin", quality)
    if quality.norm < _VAL_MIN_LINF:
        return _PreflightResult(False, "below_min_delta", quality)
    if quality.norm > eff_max:
        return _PreflightResult(False, "above_max_delta", quality)
    if quality.ssim < _VAL_MIN_SSIM:
        return _PreflightResult(False, "below_min_ssim", quality)
    if _VAL_MIN_PSNR_DB > 0.0 and quality.psnr_db < _VAL_MIN_PSNR_DB:
        return _PreflightResult(False, "below_min_psnr_db", quality)
    return _PreflightResult(True, "ok_flip", quality)


def _estimate_validator_score(
    quality: _AdvQuality,
    true_label: int,
    epsilon: float,
    response_time_ms: float,
    timeout_seconds: float,
) -> float:
    """Mirror neurons/validator.py::verify_and_score scoring formula."""
    preflight = _preflight_flip_only(quality, true_label, epsilon)
    if not preflight.ok:
        return 0.0

    effective_max_delta = min(float(epsilon), _VAL_MAX_LINF)
    denom = max(1e-12, effective_max_delta - _VAL_MIN_LINF)
    linf_ratio = min(max((quality.norm - _VAL_MIN_LINF) / denom, 0.0), 1.0)
    linf_score = (1.0 - linf_ratio) ** 2

    rmse_ratio = min(max(quality.rmse / max(1e-12, effective_max_delta), 0.0), 1.0)
    rmse_score = (1.0 - rmse_ratio) ** 2

    total_weight = max(1e-12, C.LINF_COMPONENT_WEIGHT + C.RMSE_COMPONENT_WEIGHT)
    perturbation_score = (
        (C.LINF_COMPONENT_WEIGHT * linf_score) + (C.RMSE_COMPONENT_WEIGHT * rmse_score)
    ) / total_weight

    time_ratio = response_time_ms / max(1e-12, timeout_seconds * 1000.0)
    speed_score = 1.0 - min(time_ratio, 1.0)
    return float(C.PERTURBATION_WEIGHT * perturbation_score + C.SPEED_WEIGHT * speed_score)

# ═══════════════════════════════════════════════════════════════════════════
#  Miner
# ═══════════════════════════════════════════════════════════════════════════

class PerturbMiner:
    def __init__(self, config: typing.Any) -> None:
        self.config = config
        _configure_log_level(getattr(self.config, "log_level", "DEBUG"))
        self.wallet = _make_wallet(config=self.config)
        self.subtensor = self._init_subtensor_with_retry()
        self.metagraph = self._init_metagraph_with_retry()
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

        self.model = load_efficientnet_v2_l(self.device)

        self._in_flight: dict[str, asyncio.Future] = {}
        self._in_flight_lock = asyncio.Lock()

        self.axon = _make_axon(wallet=self.wallet, config=self.config)
        self.axon.attach(
            forward_fn=self.forward,
            blacklist_fn=self.blacklist,
            priority_fn=self.priority,
        )
        self._attack_excel_path = os.path.join(
            getattr(getattr(config, "logging", None), "logging_dir", "./logs"),
            "attack_log.xlsx",
        )

    def _log_step_start(self, step_name: str, **context: typing.Any) -> None:
        if context:
            rendered = " ".join([f"{k}={v}" for k, v in context.items()])
            logger.info(f"[STEP_START] {step_name} {rendered}")
        else:
            logger.info(f"[STEP_START] {step_name}")

    def _init_subtensor_with_retry(self):
        max_attempts = int(os.getenv("SUBTENSOR_CONNECT_RETRIES", "5"))
        retry_delay_seconds = float(os.getenv("SUBTENSOR_CONNECT_RETRY_SECONDS", "4"))
        last_error = None
        for attempt in range(1, max_attempts + 1):
            try:
                logger.info(f"[MINER] Connecting subtensor (attempt {attempt}/{max_attempts})")
                return _make_subtensor(config=self.config)
            except Exception as err:
                last_error = err
                logger.warning(f"[MINER] Subtensor connect failed on attempt {attempt}: {err}")
                if attempt < max_attempts:
                    time.sleep(retry_delay_seconds * attempt)
        raise RuntimeError(f"Failed to connect subtensor after {max_attempts} attempts: {last_error}")

    def _init_metagraph_with_retry(self):
        max_attempts = int(os.getenv("METAGRAPH_SYNC_RETRIES", "5"))
        retry_delay_seconds = float(os.getenv("METAGRAPH_SYNC_RETRY_SECONDS", "4"))
        last_error = None
        for attempt in range(1, max_attempts + 1):
            try:
                logger.info(f"[MINER] Loading metagraph netuid={self.config.netuid} (attempt {attempt}/{max_attempts})")
                return self.subtensor.metagraph(netuid=self.config.netuid)
            except Exception as err:
                last_error = err
                logger.warning(f"[MINER] Metagraph load failed on attempt {attempt}: {err}")
                if attempt < max_attempts:
                    time.sleep(retry_delay_seconds * attempt)
        raise RuntimeError(f"Failed to load metagraph after {max_attempts} attempts: {last_error}")

    def sync(self) -> None:
        self.metagraph.sync(subtensor=self.subtensor)

    def _attack(
        self,
        clean: torch.Tensor,        # float C×H×W [0,1]
        true_idx: int,
        eps: float,                  # 1/255
        min_delta: float,
        deadline: float,
        log_attack: typing.Callable[..., None],
        challenge_epsilon: float,
        timeout_seconds: float,
        attack_t0: float,
    ) -> typing.Optional[dict]:
        """
        Full pipeline. Returns dict with image/k/rmse/norm/margin/pred
        or None if no flip found.
        """
        C, H, W = clean.shape
        N = C * H * W

        # ── Phase 1: Model probe ──────────────────────────────
        x = clean.detach().requires_grad_(True)
        # PREPROCESS handles resize+normalize internally
        lg = logits_for_images(self.model, x.unsqueeze(0)).squeeze(0)

        # Verify model predicts true_idx
        pred_clean = int(lg.argmax().item())
        if pred_clean != true_idx:
            logger.warning(
                f"Model predicts {pred_clean} != true_idx {true_idx} "
                f"on clean image — attacking anyway"
            )

        true_logit = lg[true_idx]
        gap = (true_logit - lg.topk(2).values[1]).item()
        logger.info(f"[PROBE] true={true_idx} gap={gap:.3f} N={N}")
        log_attack(f"[PROBE] true={true_idx} gap={gap:.3f} N={N}")
        # Top-20 candidate target classes by logit (excluding true class)
        top_classes = lg.detach().argsort(descending=True)
        top_classes = [
            c.item() for c in top_classes
            if c.item() != true_idx
        ][:20]
        logger.info(f"[TOP20CLASS] top_classes={top_classes}")

        # Compute true-label gradient once, reuse across all target rankings
        if x.grad is not None:
            x.grad.zero_()
        true_logit.backward(retain_graph=True)
        grad_true = x.grad.detach().reshape(-1).clone()
 
        # Rank targets by estimated_rmse = margin / (||g_t||2 × sqrt(N))
        targets = []
        for i, t in enumerate(top_classes):
            if x.grad is not None:
                x.grad.zero_()
            retain = (i < len(top_classes) - 1)
            lg[t].backward(retain_graph=retain)
            grad_t_raw = x.grad.detach().reshape(-1).clone()
 
            # Boundary gradient for this target
            g_t = grad_true - grad_t_raw
            margin_t = (true_logit - lg[t]).item()
            g_norm2 = g_t.norm(2).item()
            g_norm1 = g_t.abs().sum().item()
 
            if g_norm2 < 1e-10:
                continue
 
            # Feasibility: can 1/255 perturbations reach this boundary?
            if (eps * g_norm1) < margin_t * 0.1:
                continue
 
            # Sparse target estimate: more relevant for minimum changed-pixel count.
            abs_g = g_t.abs()
            n_top = max(100, int(N * 0.005))  # top 0.5%, not 1%; better sparse estimate
            top_vals = abs_g.topk(min(n_top, abs_g.numel())).values
            top_mean = float(top_vals.mean().item())
            top_sum = float(top_vals.sum().item())

            k_est = int(margin_t / max(eps * top_mean, 1e-10))

            # Lower is better. This favors classes where a few pixels have strong effect.
            sparse_score = margin_t / max(eps * top_sum, 1e-10)

            targets.append({
                "target":         t,
                "margin":         margin_t,
                "estimated_rmse": sparse_score,
                "k_estimated":    k_est,
                "gradient":       g_t,
            }) 
        if not targets:
            logger.warning("[ATTACK] No feasible targets found")
            log_attack("[ATTACK] No feasible targets found")
            return None
 
        targets.sort(key=lambda c: c["estimated_rmse"])
        target_switch_idx = 0
        best_target = targets[target_switch_idx]
        target_idx = best_target["target"]
        # Keep k_est closer to sparse estimate.
        # Over-inflating k_est makes get_batch more aggressive.
        k_est = max(int(best_target["k_estimated"] * 1.15), 300)
        logger.info(
            f"[TARGET] best={target_idx} "
            f"est_rmse={best_target['estimated_rmse']:.2e} "
            f"k_est={k_est}"
        )
        log_attack(
            f"[TARGET] best={target_idx} "
            f"est_rmse={best_target['estimated_rmse']:.2e} k_est={k_est}"
        )
        # ── Phase 2: adaptive greedy ───────────────────────────────────────
        def _estimate_gap_rate() -> typing.Optional[float]:
            """
            Estimate real gap decrease per selected pixel from recent history.
            Positive value means gap is closing.
            """
            if len(gap_history) < 2:
                return None

            window = gap_history[-6:] if len(gap_history) >= 6 else gap_history[:]
            wsum, wtot = 0.0, 0.0

            for i in range(1, len(window)):
                p0, g0, _ = window[i - 1]
                p1, g1, _ = window[i]
                dp = max(p1 - p0, 1)
                dg = g0 - g1

                if dg > 1e-9:
                    # More recent movements matter more.
                    w = 0.6 + float(i) / max(len(window) - 1, 1)
                    wsum += w * (dg / dp)
                    wtot += w

            if wtot <= 0:
                return None

            return wsum / wtot

        def _stage_limits(cur_gap: float, *, has_soft_flip: bool, phase: str) -> tuple[int, int]:
            """
            Return allowed min/max batch range for current attack stage.

            These are not the final selected batch. They define the trust-region
            search range for adaptive doubling.
            """
            if phase == "safety_grow" or has_soft_flip:
                return 1, _MAX_SAFETY_GROW_BATCH

            if cur_gap > 1.0:
                return 64, 1024
            if cur_gap > 0.1:
                return 16, 256
            if cur_gap > 0.03:
                return 8, 128
            if cur_gap > 0.01:
                return 4, 64
            if cur_gap > 0.003:
                return 2, 32

            return 1, 16

        def get_batch_estimate(k_cur, cur_gap, last_batch_time, *, has_soft_flip: bool, phase: str):
            """
            Produce only a smooth center estimate for marginal probing.

            This function should NOT decide the final batch.
            It estimates the center point around which candidate batch sizes
            will be probed by choose_batch_by_marginal_probe().
            """
            nonlocal is_first_cycle, prev_batch_size

            if is_first_cycle:
                is_first_cycle = False
                prev_batch_size = CALIBRATION_BATCH
                return CALIBRATION_BATCH, None, None

            time_remaining = grow_deadline - time.perf_counter()
            if time_remaining <= 0:
                return 1, None, None

            gap_rate = _estimate_gap_rate()

            if gap_rate is not None and gap_rate > 1e-10:
                pixels_needed = cur_gap / gap_rate
            else:
                frac_remaining = cur_gap / max(gap_at_last_switch, 1e-8)
                pixels_needed = frac_remaining * k_est

            pixels_needed = max(float(pixels_needed), 1.0)

            min_b, max_b = _stage_limits(
                cur_gap,
                has_soft_flip=has_soft_flip,
                phase=phase,
            )

            # Center estimate from real gap-rate.
            batch = int(round(pixels_needed))

            # Prevent sudden collapse before the first flip.
            # Example: 106 -> 2 should become around 37, then probe can test
            # smaller/larger candidates around that region.
            if prev_batch_size > 0 and not has_soft_flip:
                smooth_floor = int(max(1, prev_batch_size * 0.35))
                batch = max(batch, smooth_floor)

            # During safety grow, do not need a large center estimate.
            if phase == "safety_grow" or has_soft_flip:
                batch = min(batch, _MAX_SAFETY_GROW_BATCH)

            batch = max(min_b, min(batch, max_b))
            batch = min(batch, max(1, N - k_cur))
            batch = min(batch, 50000)

            return int(max(1, batch)), gap_rate, pixels_needed

        def choose_batch_by_marginal_probe(
            *,
            current_adv: torch.Tensor,
            sc: torch.Tensor,
            signs_cur: torch.Tensor,
            valid_cur: torch.Tensor,
            cur_gap: float,
            base_estimate: int,
            has_soft_flip: bool,
            phase: str,
        ) -> tuple[int, torch.Tensor]:
            """
            Adaptive doubling + trust-region refinement.

            Logic:
            1. Build stage-based doubling candidates, e.g. [64,128,256,512,1024].
            2. Probe actual gap drop for each candidate.
            3. If one candidate flips, choose the smallest flipping candidate.
            4. If larger candidate becomes weak, refine between last-good and first-weak.
               Example: 512 good, 1024 weak -> test 640, 768, 896.
            5. Choose the smallest batch that gives meaningful progress before
               marginal gain collapses.
            """
            valid_count = int(valid_cur.sum().item())
            if valid_count <= 0:
                return 0, torch.empty(0, dtype=torch.long, device=self.device)

            min_b, max_b = _stage_limits(cur_gap, has_soft_flip=has_soft_flip, phase=phase)
            max_b = min(max_b, valid_count)

            if max_b <= 0:
                return 0, torch.empty(0, dtype=torch.long, device=self.device)

            time_left = grow_deadline - time.perf_counter()

            # After soft flip, do not spend time probing. Just safety-grow with
            # a small direct batch to build robust margin.
            if (
                not _MARGINAL_PROBE_ENABLED
                or time_left < 0.45
                or phase == "safety_grow"
                or has_soft_flip
            ):
                n_pick = int(max(1, min(base_estimate, max_b)))
                top_idx = torch.topk(sc, n_pick).indices
                return n_pick, top_idx

            def _meaningful_progress_ratio(gap_value: float) -> float:
                if gap_value > 1.0:
                    return _MIN_PROGRESS_FAR
                if gap_value > 0.1:
                    return _MIN_PROGRESS_MID
                return _MIN_PROGRESS_NEAR

            def _dedupe_candidates(vals: list[int], *, cap: typing.Optional[int] = None) -> list[int]:
                """
                Deduplicate and clamp candidate batch sizes.

                Default cap is stage max_b.
                Expansion can pass cap=hard_cap to allow 1024 -> 2048 -> 4096.
                """
                limit = max_b if cap is None else int(cap)
                return sorted({
                    int(max(1, min(limit, v)))
                    for v in vals
                    if int(v) > 0
                })

            def _stage_doubling_candidates() -> list[int]:
                """
                Build initial stage-based doubling candidates.

                Stage max is only the first trust-region range.
                If the largest candidate is still strong, later logic may expand.
                """
                vals = []
                k = int(min_b)
                while k <= max_b and len(vals) < _MAX_COARSE_PROBES:
                    vals.append(k)
                    k *= 2

                return _dedupe_candidates(vals)

            def _probe_candidates(candidates: list[int]) -> tuple[list[dict], torch.Tensor]:
                """
                Probe candidate top-k groups by real model forward pass.
                """
                max_probe_k = max(candidates)
                top_pool = torch.topk(sc, max_probe_k).indices

                base_u8 = (current_adv * 255.0).round().clamp(0, 255).long()
                results = []

                with torch.no_grad():
                    lg_base = logits_for_images(
                        self.model, current_adv.unsqueeze(0)
                    ).squeeze(0).float()
                    comp_base = lg_base.clone()
                    comp_base[true_idx] = float("-inf")
                    best_other_base = int(comp_base.argmax().item())
                    base_gap = float(lg_base[true_idx].item() - lg_base[best_other_base].item())

                for k in candidates:
                    test_u8 = base_u8.clone()
                    flat_test = test_u8.reshape(-1)

                    idx = top_pool[:k]
                    s = signs_cur[idx].long()
                    flat_test[idx] = (flat_test[idx] + s).clamp(0, 255)

                    test_img = flat_test.reshape(C, H, W).float() / 255.0

                    with torch.no_grad():
                        lg_t = logits_for_images(
                            self.model, test_img.unsqueeze(0)
                        ).squeeze(0).float()
                        comp_t = lg_t.clone()
                        comp_t[true_idx] = float("-inf")
                        best_other_t = int(comp_t.argmax().item())
                        gap_t = float(lg_t[true_idx].item() - lg_t[best_other_t].item())
                        pred_t = int(lg_t.argmax().item())

                    actual_drop = base_gap - gap_t
                    efficiency = actual_drop / max(k, 1)

                    # Gradient-predicted first-order drop.
                    predicted_drop = float((eps * sc[idx].sum()).item())
                    trust_ratio = actual_drop / max(predicted_drop, 1e-9)

                    results.append({
                        "k": int(k),
                        "gap": gap_t,
                        "drop": actual_drop,
                        "eff": efficiency,
                        "trust": trust_ratio,
                        "flipped": pred_t != true_idx,
                    })

                return results, top_pool
            
            def _probe_one_k(k_probe: int, top_pool: torch.Tensor) -> dict:
                """
                Probe one exact k from the same top_pool.
                This is used to search inside coarse intervals like:
                1024..2048, 2048..4096, etc.
                """
                k_probe = int(max(1, min(k_probe, int(top_pool.numel()))))

                base_u8 = (current_adv * 255.0).round().clamp(0, 255).long()
                test_u8 = base_u8.clone()
                flat_test = test_u8.reshape(-1)

                idx = top_pool[:k_probe]
                s = signs_cur[idx].long()
                flat_test[idx] = (flat_test[idx] + s).clamp(0, 255)

                test_img = flat_test.reshape(C, H, W).float() / 255.0

                with torch.no_grad():
                    lg_t = logits_for_images(
                        self.model, test_img.unsqueeze(0)
                    ).squeeze(0).float()

                    comp_t = lg_t.clone()
                    comp_t[true_idx] = float("-inf")
                    best_other_t = int(comp_t.argmax().item())

                    gap_t = float(lg_t[true_idx].item() - lg_t[best_other_t].item())
                    pred_t = int(lg_t.argmax().item())

                return {
                    "k": k_probe,
                    "gap": gap_t,
                    "flipped": pred_t != true_idx or gap_t < 0.0,
                }


            def _fine_search_first_flip(
                *,
                top_pool: torch.Tensor,
                results: list[dict],
                chosen_flip: dict,
            ) -> tuple[int, list[dict]]:
                """
                If coarse says 2048 flips, do not blindly return 2048.
                Search inside previous non-flip -> first flip interval.

                Example:
                    1024 non-flip, 2048 flip
                    test 1152,1280,1408,1536,1664,1792,1920
                    then binary search smallest flipping k.
                """
                hi = int(chosen_flip["k"])

                lower_nonflips = [
                    int(r["k"])
                    for r in results
                    if int(r["k"]) < hi and not (r["flipped"] or r["gap"] < 0.0)
                ]
                lo = max(lower_nonflips) + 1 if lower_nonflips else 1

                dense_results = []

                interval = hi - lo + 1
                if interval >= 2048:
                    step = 128
                elif interval >= 1024:
                    step = 128
                elif interval >= 512:
                    step = 64
                elif interval >= 256:
                    step = 32
                elif interval >= 128:
                    step = 16
                elif interval >= 64:
                    step = 8
                else:
                    step = 4

                # Dense scan first: 1152,1280,1408...
                start_k = lo + step - ((lo - 1) % step)
                k = start_k

                while k < hi and (grow_deadline - time.perf_counter()) >= 0.25:
                    r = _probe_one_k(k, top_pool)
                    dense_results.append(r)
                    k += step

                dense_flips = [
                    r for r in dense_results
                    if r["flipped"] or r["gap"] < 0.0
                ]

                if dense_flips:
                    first_dense_flip = min(dense_flips, key=lambda r: r["k"])
                    hi = int(first_dense_flip["k"])

                    dense_nonflips_below = [
                        int(r["k"])
                        for r in dense_results
                        if int(r["k"]) < hi and not (r["flipped"] or r["gap"] < 0.0)
                    ]

                    if dense_nonflips_below:
                        lo = max(dense_nonflips_below) + 1

                # Binary search final interval.
                best_k = hi

                while lo <= hi and (grow_deadline - time.perf_counter()) >= 0.18:
                    mid = (lo + hi) // 2
                    r_mid = _probe_one_k(mid, top_pool)

                    if r_mid["flipped"] or r_mid["gap"] < 0.0:
                        best_k = mid
                        hi = mid - 1
                    else:
                        lo = mid + 1

                return int(best_k), dense_results

            def _find_collapse_pair(results: list[dict]) -> tuple[typing.Optional[dict], typing.Optional[dict]]:
                """
                Return (last_good, first_weak) when marginal gain collapses.
                """
                good_prev = None

                positive = [r for r in results if r["drop"] > 1e-8]
                if not positive:
                    return None, None

                good_prev = positive[0]

                for r in positive[1:]:
                    dk = max(r["k"] - good_prev["k"], 1)
                    marginal_gain = (r["drop"] - good_prev["drop"]) / dk
                    prev_eff = max(good_prev["eff"], 1e-12)
                    marginal_ratio = marginal_gain / prev_eff

                    if marginal_ratio < _TRUST_REGION_COLLAPSE_RATIO:
                        return good_prev, r

                    good_prev = r

                return good_prev, None
            
            def _largest_candidate_still_good(results: list[dict]) -> bool:
                """
                Return True when the largest tested candidate still has useful
                marginal gain, meaning the trust region should expand upward.
                """
                positive = [r for r in results if r["drop"] > 1e-8]
                if len(positive) < 2:
                    return False

                prev = positive[-2]
                last = positive[-1]

                dk = max(last["k"] - prev["k"], 1)
                marginal_gain = (last["drop"] - prev["drop"]) / dk
                prev_eff = max(prev["eff"], 1e-12)
                marginal_ratio = marginal_gain / prev_eff

                return marginal_ratio >= _TRUST_REGION_COLLAPSE_RATIO

            def _choose_from_results(results: list[dict]) -> typing.Optional[dict]:
                """
                Choose smallest flipping candidate, otherwise choose smallest batch
                with meaningful progress before marginal gain collapse.
                """
                if not results:
                    return None

                # If any candidate flips, choose the smallest flipping one.
                flipped = [r for r in results if r["flipped"] or r["gap"] < 0.0]
                if flipped:
                    return min(flipped, key=lambda r: r["k"])

                positive = [r for r in results if r["drop"] > 1e-8]
                if not positive:
                    return None

                min_progress = max(cur_gap * _meaningful_progress_ratio(cur_gap), 1e-6)

                # Prefer smallest candidate that gives meaningful absolute progress.
                meaningful = [r for r in positive if r["drop"] >= min_progress]
                if meaningful:
                    # Among meaningful candidates, avoid choosing a clearly inefficient one.
                    best_eff = max(meaningful, key=lambda r: r["eff"])
                    near_best = [
                        r for r in meaningful
                        if r["eff"] >= best_eff["eff"] * _BATCH_EFFICIENCY_TOL
                    ]
                    return min(near_best, key=lambda r: r["k"])

                # If no candidate gives meaningful progress, use marginal-collapse logic.
                last_good, first_weak = _find_collapse_pair(positive)
                if last_good is not None:
                    return last_good

                return max(positive, key=lambda r: r["drop"])

            # 1. Coarse doubling probe.
            candidates = _stage_doubling_candidates()

            if not candidates:
                n_pick = int(max(1, min(base_estimate, max_b)))
                top_idx = torch.topk(sc, n_pick).indices
                return n_pick, top_idx

            results, top_pool = _probe_candidates(candidates)

            # 1b. Expand upward if the largest candidate is still efficient.
            # This handles texture-heavy images where 1024 may still be useful.
            hard_cap = min(valid_count, _MAX_EXPANDED_BATCH)

            while (
                time_left >= _REFINE_TIME_MIN
                and candidates
                and candidates[-1] < hard_cap
                and _largest_candidate_still_good(results)
                and not any(r["flipped"] or r["gap"] < 0.0 for r in results)
            ):
                next_k = min(hard_cap, candidates[-1] * 2)
                if next_k <= candidates[-1]:
                    break

                candidates = _dedupe_candidates(candidates + [next_k], cap=hard_cap)
                results, top_pool = _probe_candidates(candidates)

                # Refresh time guard after each expansion.
                time_left = grow_deadline - time.perf_counter()

            # 2. If coarse already flips, fine-search inside the flipping interval.
            flipped = [r for r in results if r["flipped"] or r["gap"] < 0.0]
            if flipped:
                chosen = min(flipped, key=lambda r: r["k"])

                fine_k, dense_results = _fine_search_first_flip(
                    top_pool=top_pool,
                    results=results,
                    chosen_flip=chosen,
                )

                logger.info(
                    f"[BATCH_PROBE_FLIP_FINE] "
                    f"coarse={[(r['k'], round(r['gap'],5), round(r['drop'],5)) for r in results]} "
                    f"dense={[(r['k'], round(r['gap'],5), r['flipped']) for r in dense_results]} "
                    f"coarse_chosen={chosen['k']} fine_chosen={fine_k}"
                )

                return int(fine_k), top_pool[:fine_k]

            # 3. Detect trust-region collapse, then refine densely inside interval.
            last_good, first_weak = _find_collapse_pair(results)

            refine_results = []
            if (
                last_good is not None
                and first_weak is not None
                and time_left >= _REFINE_TIME_MIN
                and first_weak["k"] - last_good["k"] >= 4
                and cur_gap > 0.003
            ):
                lo = int(last_good["k"])
                hi = int(first_weak["k"])
                span = hi - lo

                # Dense interval steps.
                # Example:
                # 1024..2048 -> 1152,1280,1408,1536,1664,1792,1920
                # 512..1024  -> 576,640,704,768,832,896,960
                if span >= 2048:
                    step = 128
                elif span >= 1024:
                    step = 128
                elif span >= 512:
                    step = 64
                elif span >= 256:
                    step = 32
                elif span >= 128:
                    step = 16
                elif span >= 64:
                    step = 8
                else:
                    step = max(1, span // max(_MAX_REFINE_PROBES + 1, 2))

                refine_candidates = list(range(lo + step, hi, step))

                # Also add fractional probes as backup.
                refine_candidates += [
                    lo + int(round(span * (j / (_MAX_REFINE_PROBES + 1))))
                    for j in range(1, _MAX_REFINE_PROBES + 1)
                ]

                already = {int(r["k"]) for r in results}
                refine_candidates = sorted({
                    int(k)
                    for k in refine_candidates
                    if lo < int(k) < hi and int(k) not in already
                })

                if refine_candidates:
                    all_candidates = _dedupe_candidates(
                        candidates + refine_candidates,
                        cap=min(valid_count, _MAX_EXPANDED_BATCH),
                    )

                    results, top_pool = _probe_candidates(all_candidates)
                    refine_results = [r for r in results if r["k"] in refine_candidates]

                    logger.info(
                        f"[BATCH_REFINE_DENSE] lo={lo} hi={hi} "
                        f"refine={[(r['k'], round(r['gap'],5), round(r['drop'],5)) for r in refine_results]}"
                    )

                    # If dense refine discovers a flip, do not return the coarse flip.
                    # Fine-search the smallest flipping k immediately.
                    flipped_after_refine = [
                        r for r in results
                        if r["flipped"] or r["gap"] < 0.0
                    ]

                    if flipped_after_refine:
                        chosen_flip = min(flipped_after_refine, key=lambda r: r["k"])

                        fine_k, dense_flip_results = _fine_search_first_flip(
                            top_pool=top_pool,
                            results=results,
                            chosen_flip=chosen_flip,
                        )

                        logger.info(
                            f"[BATCH_REFINE_FLIP_FINE] "
                            f"chosen_flip={chosen_flip['k']} fine_chosen={fine_k} "
                            f"dense={[(r['k'], round(r['gap'],5), r['flipped']) for r in dense_flip_results]}"
                        )

                        return int(fine_k), top_pool[:fine_k]

            # 4. Choose final candidate.
            chosen = _choose_from_results(results)

            if chosen is None:
                # Gradient is stale or target is bad. Use small stage-minimum batch.
                n_pick = int(max(1, min(min_b, max_b)))
                top_idx = torch.topk(sc, n_pick).indices
                logger.info(
                    f"[BATCH_PROBE_BAD] no positive drop; fallback={n_pick} "
                    f"candidates={candidates}"
                )
                return n_pick, top_idx

            n_pick = int(max(1, min(chosen["k"], valid_count, _MAX_EXPANDED_BATCH)))

            logger.info(
                f"[BATCH_PROBE_TRUST] coarse={[(r['k'], round(r['gap'],5), round(r['drop'],5)) for r in results if r['k'] in candidates]} "
                f"refine={[(r['k'], round(r['gap'],5), round(r['drop'],5)) for r in refine_results]} "
                f"chosen={n_pick} gap_after={chosen['gap']:.5f} "
                f"drop={chosen['drop']:.5e} eff={chosen['eff']:.2e} "
                f"trust={chosen['trust']:.2f}"
            )

            return n_pick, top_pool[:n_pick]
        def check_flip(adv_float):
            """
            Verify flip on uint8-snapped + PNG-roundtripped image to match validator numerics.
            margin = true_logit - best_non_true_logit.
            margin < 0  ⇒ soft flip.
            margin <= _FLIP_MARGIN_EPS ⇒ hard validator-safe flip.
            """
            snapped_u8 = (adv_float * 255.0).round().clamp(0, 255) / 255.0
            try:
                snapped = _encode_decode_roundtrip(snapped_u8)
            except Exception:
                snapped = snapped_u8

            with torch.no_grad():
                lg_s = logits_for_images(
                    self.model, snapped.unsqueeze(0)
                ).squeeze(0).float()

                pred = int(lg_s.argmax().item())
                comp = lg_s.clone()
                comp[true_idx] = float("-inf")
                best_other = int(comp.argmax().item())
                margin = float(lg_s[true_idx].item() - lg_s[best_other].item())

            return pred != true_idx, margin, snapped, pred

        def _is_soft_flip(margin: float) -> bool:
            return margin < 0.0

        def _is_hard_flip(margin: float) -> bool:
            return margin <= _FLIP_MARGIN_EPS

        def _is_strong_safe(margin: float) -> bool:
            return margin <= _STRONG_SAFE_MARGIN

        def _record_candidate(
            snapped,
            margin: float,
            pred: int,
            *,
            tag: str,
            gap_now: float,
            selected_snapshot: typing.Optional[list[int]] = None,
        ) -> None:
            nonlocal best_hard_result, best_soft_result, best_progress_result, best_progress_gap

            sel_snapshot = selected.copy() if selected_snapshot is None else selected_snapshot.copy()
            if len(sel_snapshot) == 0:
                return

            flip_q = _measure_adv_quality(
                self.model, clean, snapped, true_idx, true_idx
            )
            cand = {
                "image":    snapped.clone(),
                "k":        len(sel_snapshot),
                "margin":   margin,
                "selected": sel_snapshot,
                "pred":     pred,
                "rmse_val": flip_q.rmse,
                "norm_val": flip_q.norm,
            }
            if gap_now < best_progress_gap:
                best_progress_gap = gap_now
                best_progress_result = cand
            if _is_soft_flip(margin):
                if (
                    best_soft_result is None
                    or flip_q.rmse < best_soft_result.get("rmse_val", float("inf"))
                ):
                    best_soft_result = cand
                    log_attack(
                        f"[{tag}_SOFT] k={len(sel_snapshot)} margin={margin:.5f}",
                        prediction=pred,
                        rmse=flip_q.rmse,
                        norm=flip_q.norm,
                    )
            if _is_hard_flip(margin):
                preflight = _preflight_flip_only(flip_q, true_idx, challenge_epsilon)
                if preflight.ok and (
                    best_hard_result is None
                    or flip_q.rmse < best_hard_result.get("rmse_val", float("inf"))
                ):
                    best_hard_result = cand
                    log_attack(
                        f"[{tag}_HARD] k={len(sel_snapshot)} margin={margin:.5f}",
                        prediction=pred,
                        rmse=flip_q.rmse,
                        norm=flip_q.norm,
                        estimated_score=_estimate_validator_score(
                            flip_q,
                            true_idx,
                            challenge_epsilon,
                            (time.perf_counter() - attack_t0) * 1000.0,
                            timeout_seconds,
                        ),
                    )

        current_adv = clean.clone()
        selected = []                                        # pixel indices
        # selected_mask = torch.zeros(N, dtype=torch.bool, device=self.device)
        selected_signs = torch.zeros(N, dtype=torch.long, device=self.device)
        best_hard_result = None
        best_soft_result = None
        best_progress_result = None
        best_progress_gap = float("inf")
        attack_phase = "grow"   # grow | safety_grow
        near_boundary_steps = 0
        has_soft_flip = False
        gap_initial = gap
        cur_gap = gap
        gap_at_last_switch = gap  # tracks gap_initial for fallback after target switches
        g_t = best_target["gradient"]
        g_norm1 = g_t.abs().sum().item()
        lambda_prior = (eps * g_norm1) / (gap_initial * N)

        CALIBRATION_BATCH = max(5, int(0.10 / max(lambda_prior, 1e-8)))
        CALIBRATION_BATCH = min(CALIBRATION_BATCH, 200)
        is_first_cycle = True

        gap_history = [(0, gap_initial, time.perf_counter())]
        ELIM_RESERVE = float(os.getenv("MINER_ELIM_RESERVE", "2.6"))
        FINAL_BUFFER = 0.25
        grow_deadline = deadline - ELIM_RESERVE
        final_deadline = deadline - FINAL_BUFFER
        last_batch_time = 0.20
        prev_batch_size = CALIBRATION_BATCH

        while True:
            # Stop growing early. Remaining time is for elimination.
            if time.perf_counter() > grow_deadline:
                break
            cycle_start = time.perf_counter()
            x_in = current_adv.detach().requires_grad_(True)
            lg_in = logits_for_images(self.model, x_in.unsqueeze(0)).squeeze(0)

            # Compute gaps from the SAME forward pass — no extra forward needed.
            with torch.no_grad():
                comp = lg_in.detach().clone()
                comp[true_idx] = float("-inf")
                runner_up_idx = int(comp.argmax().item())

                # Actual untargeted flip boundary.
                rank2_gap = float(lg_in[true_idx].item() - lg_in[runner_up_idx].item())

                # Selected target boundary used by the current gradient.
                target_gap = float(lg_in[true_idx].item() - lg_in[target_idx].item())

            # Dynamic target switch: follow runner-up when it is clearly closer.
            if runner_up_idx != target_idx and rank2_gap + 0.02 < target_gap:
                old_target_idx = target_idx
                target_idx = runner_up_idx
                target_gap = rank2_gap
                cur_gap = rank2_gap
                logger.info(
                    f"[TARGET_SWITCH] old={old_target_idx} new={target_idx} "
                    f"old_target_gap={float(lg_in[true_idx].item() - lg_in[old_target_idx].item()):.4f} "
                    f"new_target_gap={target_gap:.4f} rank2_gap={rank2_gap:.4f}"
                )
                matching_target = next(
                    (t for t in targets if int(t["target"]) == int(target_idx)),
                    None,
                )
                if matching_target is not None:
                    k_est = max(int(matching_target.get("k_estimated", k_est) * 1.15), 300)
                else:
                    # Runner-up was not in the pre-ranked target list.
                    # Keep estimate conservative and avoid premature stuck-break.
                    k_est = max(int(k_est * 0.75), len(selected) + 300, 300)

                logger.info(f"[TARGET_SWITCH_K_EST] target={target_idx} k_est_updated={k_est}")

                gap_history = [(len(selected), max(cur_gap, 1e-6), time.perf_counter())]
                gap_at_last_switch = cur_gap
                near_boundary_steps = 0
            else:
                cur_gap = min(target_gap, rank2_gap)

            # Near-boundary backup: try runner-up targets when stuck close to boundary.
            if cur_gap < _NEAR_BOUNDARY_GAP and not has_soft_flip:
                near_boundary_steps += 1
                if (
                    near_boundary_steps >= _TARGET_SWITCH_STEPS
                    and target_switch_idx < min(2, len(targets) - 1)
                ):
                    target_switch_idx += 1
                    alt = targets[target_switch_idx]
                    target_idx = alt["target"]
                    g_t = alt["gradient"]
                    g_norm1 = g_t.abs().sum().item()
                    k_est = max(int(alt.get("k_estimated", k_est) * 1.15), 300)

                    target_gap = float(lg_in[true_idx].item() - lg_in[target_idx].item())
                    cur_gap = min(target_gap, rank2_gap)
                    gap_history = [(len(selected), max(cur_gap, 1e-6), time.perf_counter())]
                    gap_at_last_switch = cur_gap
                    near_boundary_steps = 0
                    logger.info(
                        f"[TARGET_BACKUP] idx={target_switch_idx} target={target_idx} "
                        f"gap={cur_gap:.4f} k_est={k_est}"
                    )
                    log_attack(
                        f"[TARGET_BACKUP] target={target_idx} gap={cur_gap:.4f} k_est={k_est}"
                    )
            else:
                near_boundary_steps = 0

            logger.info(
                f"[ATTACK] phase={attack_phase} target_gap={target_gap:.4f} "
                f"rank2_gap={rank2_gap:.4f} target_idx={target_idx} "
                f"runner_up_idx={runner_up_idx} cur_gap={cur_gap:.4f}"
            )

            # Two backward passes sharing the same retained graph.
            # IMPORTANT: this now uses the possibly switched target_idx.
            if x_in.grad is not None:
                x_in.grad.zero_()

            lg_in[true_idx].backward(retain_graph=True)
            g_true_cur = x_in.grad.detach().reshape(-1).clone()

            x_in.grad.zero_()
            lg_in[target_idx].backward()
            g_tgt_cur = x_in.grad.detach().reshape(-1).clone()

            g_cur = g_true_cur - g_tgt_cur
            
            gap_history.append((len(selected), max(cur_gap, 1e-6), time.perf_counter()))
            if len(gap_history) > 12:
                gap_history.pop(0)

            # With direct large batches, k_est is frequently exceeded early.
            # Only break if gap is genuinely stuck (not decreasing over last 4 cycles).
            if len(selected) > k_est * 4.0 and best_hard_result is None and cur_gap > 0.05:
                gap_stuck = False
                if len(gap_history) >= 4:
                    recent_gaps = [g for _, g, _ in gap_history[-4:]]
                    total_recent_drop = recent_gaps[0] - recent_gaps[-1]
                    gap_stuck = total_recent_drop < cur_gap * 0.02

                if gap_stuck:
                    logger.warning(
                        f"[ATTACK] Exceeded 4x k_est={k_est} at k={len(selected)} "
                        f"gap={cur_gap:.4f} gap_stuck=True — switching target or breaking"
                    )
                    if runner_up_idx != target_idx:
                        target_idx = runner_up_idx
                        cur_gap = rank2_gap
                        gap_history = [(len(selected), max(cur_gap, 1e-6), time.perf_counter())]
                        continue
                    break

            # Periodic flip check: record candidates, transition phases.
            if cur_gap < 2.0 or rank2_gap < 0.05:
                _flipped, _margin, _snapped, _pred = check_flip(current_adv)
                if _flipped:
                    has_soft_flip = True
                    if attack_phase == "grow":
                        attack_phase = "safety_grow"
                        logger.info(
                            f"[FIRST_SOFT_FLIP] k={len(selected)} margin={_margin:.5f} "
                            f"— entering safety grow"
                        )
                        log_attack(
                            f"[FIRST_SOFT_FLIP] k={len(selected)} margin={_margin:.5f}"
                        )
                    _record_candidate(
                        _snapped, _margin, _pred,
                        tag="CHECK", gap_now=min(cur_gap, rank2_gap), selected_snapshot=selected.copy(),
                    )
                    # Stop growth as soon as PNG-roundtrip hard flip is barely safe.
                    # Do not keep growing to deep safety margin.
                    if _is_hard_flip(_margin) and best_hard_result is not None:
                        logger.info(
                            f"[BOUNDARY_SAFE_CHECK] k={len(selected)} "
                            f"margin={_margin:.6f} — stopping growth"
                        )
                        break

            # First get only a smooth estimate. Final batch will be selected
            # after current valid pixels and saliency are known.
            batch_estimate, _, _ = get_batch_estimate(
                len(selected), cur_gap, last_batch_time,
                has_soft_flip=has_soft_flip, phase=attack_phase,
            )

            # Scores for this step
            flat_cur = current_adv.reshape(-1)
            flat_clean = clean.reshape(-1)
            sc = g_cur.abs()

            # Direction from current gradient.
            signs_cur = -g_cur.sign()
            signs_cur[signs_cur == 0] = 1.0

            # Current uint8 and clean uint8 are the source of truth.
            q_cur = (flat_cur * 255.0).round().clamp(0, 255)
            q_clean = (flat_clean * 255.0).round().clamp(0, 255)

            # IMPORTANT:
            # Only allow pixels still equal to clean.
            # This prevents +2/255 or -2/255 even if selected_mask is wrong.
            unchanged_from_clean = (q_cur == q_clean)

            # After applying one step, pixel must remain valid.
            valid_step = (
                ((signs_cur > 0) & (q_cur < 255)) |
                ((signs_cur < 0) & (q_cur > 0))
            )

            valid_cur = unchanged_from_clean & valid_step

            sc[~valid_cur] = 0.0
 
            if sc.max() <= 0:
                break
            
            n_pick, top_idx = choose_batch_by_marginal_probe(
                current_adv=current_adv,
                sc=sc,
                signs_cur=signs_cur,
                valid_cur=valid_cur,
                cur_gap=cur_gap,
                base_estimate=batch_estimate,
                has_soft_flip=has_soft_flip,
                phase=attack_phase,
            )

            if n_pick <= 0 or top_idx.numel() == 0:
                break

            # Vectorized — replaces entire for loop
            s_vec = signs_cur[top_idx].long()          # signs for selected pixels
            adv_u8 = (current_adv * 255.0).round().clamp(0, 255).long()
            flat_u8 = adv_u8.reshape(-1)
            flat_u8[top_idx] = (flat_u8[top_idx] + s_vec).clamp(0, 255)
            current_adv = flat_u8.reshape(C, H, W).float() / 255.0

            # Update selected list and signs.
            # selected_mask removed: unchanged_from_clean already prevents re-selection.
            new_indices = top_idx.tolist()
            selected.extend(new_indices)
            selected_signs[top_idx] = s_vec
            last_batch_time = 0.7 * last_batch_time + 0.3 * max(time.perf_counter() - cycle_start, 0.02)
            prev_batch_size = n_pick
            # Post-batch flip check: record and decide whether to stop growth.
            _flipped2, _margin2, _snapped2, _pred2 = check_flip(current_adv)

            if len(selected) > 0:
                _record_candidate(
                    _snapped2,
                    _margin2,
                    _pred2,
                    tag="BATCH",
                    gap_now=_margin2,
                    selected_snapshot=selected.copy(),
                )

            if _flipped2:
                has_soft_flip = True

                if attack_phase == "grow":
                    attack_phase = "safety_grow"
                    logger.info(
                        f"[FIRST_FLIP] k={len(selected)} margin={_margin2:.6f} "
                        f"— entering tiny safety mode"
                    )
                    log_attack(
                        f"[FIRST_FLIP] k={len(selected)} margin={_margin2:.6f}"
                    )

                # Main change:
                # Stop as soon as the PNG-roundtrip hard flip is barely safe.
                # Do not push to deep negative margin.
                if _is_hard_flip(_margin2):
                    logger.info(
                        f"[BOUNDARY_SAFE_FLIP] k={len(selected)} margin={_margin2:.6f} "
                        f"— stop growth and start elimination"
                    )
                    log_attack(
                        f"[BOUNDARY_SAFE_FLIP] k={len(selected)} margin={_margin2:.6f}"
                    )
                    break

                # Soft flip only. Continue with tiny safety batch.
                # Because _MAX_SAFETY_GROW_BATCH is now 4, this will not explode RMSE.
                logger.info(
                    f"[SOFT_FLIP_TINY_GROW] k={len(selected)} margin={_margin2:.6f} "
                    f"needs <= {_FLIP_MARGIN_EPS:.6f}"
                )
                continue

        # Resolve best candidate: hard > soft > lowest-gap progress > None.
        best_result = best_hard_result
        if best_result is None:
            _flipped_final, _margin_final, _snapped_final, _pred_final = check_flip(current_adv)
            if len(selected) > 0:
                _record_candidate(
                    _snapped_final, _margin_final, _pred_final,
                    tag="LATE", gap_now=min(cur_gap, rank2_gap), selected_snapshot=selected.copy(),
                )
            best_result = best_hard_result

        # Do NOT submit soft/progress candidates.
        # Validator rejects these as label_match_with_original when the margin is too close to zero.
        if best_result is None:
            logger.warning(
                "[NO_HARD_FLIP] No validator-safe hard flip found; returning None instead of soft/progress candidate"
            )
            log_attack("[NO_HARD_FLIP] no validator-safe hard flip")
            return None

        if best_result is None:
            if len(selected) == 0:
                logger.warning("[ATTACK] No perturbation created — returning clean")
                log_attack("[ATTACK] No perturbation created")
                return None
            logger.warning(
                f"[ATTACK] No hard flip after {len(selected)} pixels"
            )
            log_attack(f"[ATTACK] No hard flip after {len(selected)} pixels")
            return None

        # ── Phase 4: backward elimination ─────────────────────────────────
        # Use all reserved time for elimination.
        elim_budget = max(0.0, final_deadline - time.perf_counter())
        elim_deadline = final_deadline

        logger.info(f"[ELIM] k_initial={best_result['k']} elim_budget={elim_budget:.3f}")
        sel = best_result["selected"].copy()
        curr = best_result["image"].clone()

        def rebuild(sel_list):
            base = (clean * 255.0).round().clamp(0, 255).long()
            flat_b = base.reshape(-1)
            if sel_list:
                sel_t = torch.tensor(sel_list, dtype=torch.long, device=self.device)
                s = selected_signs[sel_t]
                flat_b[sel_t] = (flat_b[sel_t] + s).clamp(0, 255)
            return flat_b.reshape(C, H, W).float() / 255.0

        def _probe_margin(adv_float) -> typing.Optional[float]:
            """Return margin if hard-flip gates pass, else None."""
            try:
                decoded = _encode_decode_roundtrip(adv_float)
            except Exception:
                decoded = (adv_float * 255.0).round().clamp(0, 255) / 255.0

            diff = decoded - clean
            norm = float(diff.abs().max().item())
            eff_max = min(float(challenge_epsilon), _VAL_MAX_LINF)
            if norm < _VAL_MIN_LINF or norm > eff_max:
                return None

            with torch.inference_mode():
                logits = _model_logits_batch(
                    self.model, decoded.unsqueeze(0).float()
                ).squeeze(0).float()
            if int(logits.argmax().item()) == true_idx:
                return None
            competitor = logits.clone()
            competitor[true_idx] = float("-inf")
            margin = float(logits[true_idx].item() - competitor.max().item())
            if margin > _FLIP_MARGIN_EPS:
                return None
            return margin

        def still_flips(adv_float) -> bool:
            # Margin-preserving: only accept if hard flip survives PNG roundtrip.
            return _probe_margin(adv_float) is not None

        # ── Prefix binary shrink first ─────────────────────────────
        # Finds smallest prefix of selected pixels that still flips.
        # This is very fast and often removes a lot of late-added pixels.
        lo, hi = 1, len(sel)
        best_prefix = sel
        best_prefix_adv = curr

        while lo <= hi and time.perf_counter() < elim_deadline:
            mid = (lo + hi) // 2
            trial_sel = sel[:mid]
            trial_adv = rebuild(trial_sel)

            if still_flips(trial_adv):
                best_prefix = trial_sel
                best_prefix_adv = trial_adv
                hi = mid - 1
            else:
                lo = mid + 1

        if len(best_prefix) < len(sel):
            logger.info(
                f"[ELIM_PREFIX] k {len(sel)} -> {len(best_prefix)}"
            )
            sel = best_prefix
            curr = best_prefix_adv

        block = max(1, len(sel) // 4)
        while block >= 1 and time.perf_counter() < elim_deadline:
            improved = False
            # Try removing lower-priority later pixels first.
            starts = list(range(max(0, len(sel) - block), -1, -block))
            if 0 not in starts:
                starts.append(0)

            for start in starts:
                if time.perf_counter() >= elim_deadline:
                    break

                end = min(start + block, len(sel))
                if end <= start:
                    continue

                trial_sel = sel[:start] + sel[end:]
                if not trial_sel:
                    continue

                trial = rebuild(trial_sel)

                if still_flips(trial):
                    sel = trial_sel
                    curr = trial
                    improved = True
                    logger.info(
                        f"[ELIM_CHUNK] removed={end-start} k={len(sel)} block={block}"
                    )
                    break

            if not improved:
                block //= 2
        improved = True
        logger.info(f"[ELIM] k_initial={len(sel)} elim_budget={elim_budget}")

        while improved and time.perf_counter() < elim_deadline:
            improved = False
 
            # Rank pixels by weakest gradient contribution — try removing first
            x_e = curr.detach().requires_grad_(True)
            lg_e = logits_for_images(
                self.model, x_e.unsqueeze(0)
            ).squeeze(0)
            lg_e[true_idx].backward()
            g_e = x_e.grad.detach().reshape(-1).abs()
            order = sorted(
                range(len(sel)),
                key=lambda i: g_e[sel[i]].item()
            )
 
            for pos in order:
                if time.perf_counter() > elim_deadline:
                    break
                trial = sel[:pos] + sel[pos + 1:]
                trial_adv = rebuild(trial)
                if still_flips(trial_adv):
                    sel = trial
                    curr = trial_adv
                    improved = True
                    break   # restart with updated gradient
 
        k_final = len(sel)
        logger.info(f"[ELIM] k_final={k_final}")

        _, final_q = _quality_on_png(self.model, clean, curr, true_idx, true_idx)
        final_score = _estimate_validator_score(
            final_q,
            true_idx,
            challenge_epsilon,
            (time.perf_counter() - attack_t0) * 1000.0,
            timeout_seconds,
        )

        final_pred = int(final_q.pred)
        final_margin = float(final_q.margin)

        # Safety: if elimination weakened margin below hard-flip threshold, revert.
        final_preflight = _preflight_flip_only(final_q, true_idx, challenge_epsilon)
        if not final_preflight.ok:
            logger.warning(
                f"[FINAL_INVALID_AFTER_ELIM] reason={final_preflight.reason} "
                f"margin={final_margin:.5f} k_elim={k_final}; "
                f"falling back to pre-elim hard flip k={best_result['k']} "
                f"margin={best_result['margin']:.5f}"
            )
            curr = best_result["image"].clone()
            k_final = int(best_result["k"])
            _, final_q = _quality_on_png(self.model, clean, curr, true_idx, true_idx)
            final_score = _estimate_validator_score(
                final_q,
                true_idx,
                challenge_epsilon,
                (time.perf_counter() - attack_t0) * 1000.0,
                timeout_seconds,
            )
            final_pred = int(final_q.pred)
            final_margin = float(final_q.margin)

        rmse = float(final_q.rmse)

        logger.info(
            f"[FINAL] k={k_final} rmse={rmse:.2e} "
            f"norm={final_q.norm:.4f} margin={final_margin:.5f} "
            f"pred={final_pred} est_score={final_score:.4f}"
        )
        log_attack(
            f"[FINAL] k={k_final} margin={final_margin:.3f}",
            prediction=final_pred,
            rmse=final_q.rmse,
            norm=final_q.norm,
            estimated_score=final_score,
        )
        final_preflight = _preflight_flip_only(final_q, true_idx, challenge_epsilon)
        if not final_preflight.ok:
            logger.warning(
                f"[FINAL_BLOCKED] reason={final_preflight.reason} "
                f"pred={final_q.pred} margin={final_q.margin:.5f} "
                f"rmse={final_q.rmse:.2e} norm={final_q.norm:.6f}"
            )
            log_attack(
                f"[FINAL_BLOCKED] reason={final_preflight.reason} margin={final_q.margin:.5f}",
                prediction=final_q.pred,
                rmse=final_q.rmse,
                norm=final_q.norm,
                estimated_score=0.0,
            )
            return None
        return {
            "image":  curr,
            "k":      k_final,
            "rmse":   final_q.rmse,
            "norm":   final_q.norm,
            "margin": final_margin,
            "pred":   final_pred,
        }

    async def forward(self, synapse: AttackChallenge) -> AttackChallenge:
        self._log_step_start(
            "miner_forward",
            task_id=getattr(synapse, "task_id", "unknown"),
            norm_type=getattr(synapse, "norm_type", "unknown"),
            epsilon=getattr(synapse, "epsilon", "unknown"),
        )

        task_id = getattr(synapse, "task_id", "unknown")
        model_name = getattr(synapse, "model_name", "")
        prompt = getattr(synapse, "prompt", "")
        true_label = getattr(synapse, "true_label", "")
        synapse_epsilon = float(getattr(synapse, "epsilon", 0.0))
        norm_type = getattr(synapse, "norm_type", "")
        min_delta = float(getattr(synapse, "min_delta", 1.0 / 255.0))

        resolution = "unknown"
        attack_log = _AttackExcelRecorder(
            self._attack_excel_path,
            task_id=task_id,
            model_name=model_name,
            prompt=prompt,
            true_label=true_label,
            epsilon=synapse_epsilon,
            norm_type=norm_type,
            min_delta=min_delta,
            resolution=resolution,
        )

        def log_attack(
            progress: str,
            prediction: typing.Optional[int] = None,
            rmse: typing.Optional[float] = None,
            norm: typing.Optional[float] = None,
            estimated_score: typing.Optional[float] = None,
        ) -> None:
            attack_log.log(
                progress,
                prediction=prediction,
                rmse=rmse,
                norm=norm,
                estimated_score=estimated_score,
            )

        try:
            if synapse.norm_type != "Linf":
                logger.info(f"Skipping task={task_id}: unsupported norm_type={synapse.norm_type}")
                log_attack(f"[SKIP] unsupported norm_type={synapse.norm_type}")
                synapse.perturbed_image_b64 = synapse.clean_image_b64
                return synapse

            clean = decode_image_b64(synapse.clean_image_b64).to(self.device)
            true_idx = resolve_target_index(synapse.true_label)
            c, h, w = clean.shape
            resolution = f"{c}x{h}x{w}"
            attack_log.set_resolution(resolution)
            log_attack("[START] attack forward")

            if true_idx is None:
                logger.warning(
                    f"Skipping task={task_id}: unresolved true_label={getattr(synapse, 'true_label', None)}"
                )
                log_attack(f"[SKIP] unresolved true_label={true_label}")
                synapse.perturbed_image_b64 = synapse.clean_image_b64
                return synapse

            epsilon = 1.0/255.0
            budget = 13.0
            t0 = time.perf_counter()
            deadline = t0 + budget
            timeout_seconds = float(getattr(synapse, "timeout_seconds", C.TIMEOUT_SECONDS))
            logger.info(
                f"[FORWARD] task_eps={epsilon:.4f} res={c}x{h}x{w} "
                f"val_linf=[{_VAL_MIN_LINF:.4f},{min(epsilon, _VAL_MAX_LINF):.4f}]"
            )
            try:
                result = self._attack(
                    clean,
                    true_idx,
                    epsilon,
                    min_delta,
                    deadline,
                    log_attack,
                    challenge_epsilon=synapse_epsilon,
                    timeout_seconds=timeout_seconds,
                    attack_t0=t0,
                )
            except Exception as e:
                logger.warning(f"Attack failed with exception: {e}")
                log_attack(f"[ERROR] attack exception: {e}")
                result = None
            if result is None:
                logger.warning("No perturbation produced — returning clean image")
                log_attack("[DONE] no perturbation — returning clean image", estimated_score=0.0)
                synapse.perturbed_image_b64 = synapse.clean_image_b64
                return synapse

            synapse.perturbed_image_b64 = encode_image_b64(result["image"])
            _, submit_q = _quality_on_png(
                self.model, clean, result["image"], true_idx, true_idx
            )
            submit_score = _estimate_validator_score(
                submit_q,
                true_idx,
                synapse_epsilon,
                (time.perf_counter() - t0) * 1000.0,
                timeout_seconds,
            )
            logger.info(
                f"Finished task={task_id} "
                f"true={true_idx} pred={result['pred']} "
                f"k={result['k']} rmse={submit_q.rmse:.2e} "
                f"norm={submit_q.norm:.4f} margin={result['margin']:.3f} "
                f"est_score={submit_score:.4f}"
            )
            log_attack(
                f"[DONE] submitted k={result['k']} margin={result['margin']:.3f}",
                prediction=result["pred"],
                rmse=submit_q.rmse,
                norm=submit_q.norm,
                estimated_score=submit_score,
            )
            return synapse
        finally:
            attack_log.flush()

    async def blacklist(self, synapse: AttackChallenge) -> typing.Tuple[bool, str]:
        self._log_step_start(
            "miner_blacklist",
            task_id=getattr(synapse, "task_id", "unknown"),
            caller_hotkey=getattr(getattr(synapse, "dendrite", None), "hotkey", None),
        )
        if synapse.dendrite is None or synapse.dendrite.hotkey is None:
            logger.warning("Blacklist reject: missing caller hotkey")
            return True, "Missing caller hotkey"

        hotkey = synapse.dendrite.hotkey
        if hotkey not in self.metagraph.hotkeys:
            logger.warning(f"Blacklist reject: unregistered caller hotkey={hotkey}")
            return True, "Unregistered caller"

        uid = self.metagraph.hotkeys.index(hotkey)
        if not self.metagraph.validator_permit[uid]:
            logger.warning(f"Blacklist reject: caller uid={uid} lacks validator permit")
            return True, "Caller is not validator"

        logger.info(f"Blacklist allow: caller uid={uid} hotkey={hotkey}")
        return False, "OK"

    async def priority(self, synapse: AttackChallenge) -> float:
        self._log_step_start(
            "miner_priority",
            task_id=getattr(synapse, "task_id", "unknown"),
            caller_hotkey=getattr(getattr(synapse, "dendrite", None), "hotkey", None),
        )
        if synapse.dendrite is None or synapse.dendrite.hotkey is None:
            logger.info("Priority=0.0: missing caller hotkey")
            return 0.0
        if synapse.dendrite.hotkey not in self.metagraph.hotkeys:
            logger.info(f"Priority=0.0: unknown hotkey={synapse.dendrite.hotkey}")
            return 0.0
        uid = self.metagraph.hotkeys.index(synapse.dendrite.hotkey)
        priority = float(self.metagraph.S[uid])
        logger.info(f"Priority computed: uid={uid} priority={priority:.6f}")
        return priority

    def run(self) -> None:
        self.sync()

        if self.wallet.hotkey.ss58_address not in self.metagraph.hotkeys:
            raise RuntimeError("Miner hotkey is not registered on this netuid.")

        logger.info(
            f"Serving miner axon {self.axon} on network: {self.config.subtensor.network} with netuid: {self.config.netuid}"
        )
        self.axon.serve(netuid=self.config.netuid, subtensor=self.subtensor)
        self.axon.start()

        logger.info("Miner started. Waiting for validator queries.")
        while True:
            time.sleep(12)
            self.sync()


def build_config() -> typing.Any:
    parser = argparse.ArgumentParser(description="Perturb subnet miner (default baseline)")
    parser.add_argument("--netuid", type=int, required=True)
    parser.add_argument("--network", type=str, default=os.getenv("NETWORK", "finney"))
    parser.add_argument(
        "--subtensor.chain_endpoint",
        dest="chain_endpoint",
        type=str,
        default=os.getenv("SUBTENSOR_CHAIN_ENDPOINT", os.getenv("CHAIN_ENDPOINT", "")),
    )
    parser.add_argument("--wallet.name", dest="wallet_name", type=str, default=os.getenv("WALLET_NAME", "default"))
    parser.add_argument("--wallet.hotkey", dest="wallet_hotkey", type=str, default=os.getenv("HOTKEY_NAME", "default"))
    parser.add_argument("--logging-dir", dest="logging_dir", type=str, default=os.getenv("LOGGING_DIR", "./logs"))
    parser.add_argument("--log-level", dest="log_level", type=str, default=os.getenv("LOG_LEVEL", "DEBUG"))
    parser.add_argument(
        "--axon.port",
        dest="axon_port",
        type=int,
        default=int(os.getenv("MINER_PORT", os.getenv("AXON_PORT", "9000"))),
    )

    if hasattr(bt, "config"):
        config = bt.config(parser)
    else:
        config = parser.parse_args()

    if not hasattr(config, "wallet"):
        config.wallet = type("WalletConfig", (), {})()
    config.wallet.name = getattr(config.wallet, "name", getattr(config, "wallet_name", "default"))
    config.wallet.hotkey = getattr(config.wallet, "hotkey", getattr(config, "wallet_hotkey", "default"))

    if not hasattr(config, "subtensor"):
        config.subtensor = type("SubtensorConfig", (), {})()
    config.subtensor.network = getattr(config.subtensor, "network", getattr(config, "network", "finney"))
    config.subtensor.chain_endpoint = getattr(
        config.subtensor, "chain_endpoint", getattr(config, "chain_endpoint", "")
    )

    if not hasattr(config, "logging"):
        config.logging = type("LoggingConfig", (), {})()
    config.logging.logging_dir = getattr(config.logging, "logging_dir", getattr(config, "logging_dir", "./logs"))

    if not hasattr(config, "axon"):
        config.axon = type("AxonConfig", (), {})()
    config.axon.port = int(getattr(config.axon, "port", getattr(config, "axon_port", 9000)))

    config.log_level = getattr(config, "log_level", os.getenv("LOG_LEVEL", "DEBUG"))

    return config


if __name__ == "__main__":
    miner = PerturbMiner(config=build_config())
    miner.run()

